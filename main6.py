import pdfplumber
import re
import json
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def clean_narration(text):
    """
    Extract clean product name:
    - Remove content inside parentheses (SKU, size, color codes)
    - Remove HSN lines
    - Remove underscored variant codes
    - Strip extra whitespace
    """
    # Remove HSN lines
    text = re.sub(r'\bHSN:\d+\b', '', text)

    # Remove anything inside parentheses
    text = re.sub(r'\(.*?\)', '', text)

    # Take only the part before the pipe that precedes the ASIN (B0...).
    # Some products use pipes as name separators, e.g.:
    #   "Puma | Essentials Hoodie | Black | S | B0D2R5MWP5"
    # so we must cut at the LAST pipe before the ASIN, not the first pipe.
    # Cut at the pipe that directly precedes the ASIN (B0...).
    # Pattern: "Name parts | possibly | color | size | B0XXXXXXXX"
    # Strategy: find the ASIN, then walk backwards through pipe-segments,
    # dropping short segments (colours/sizes, typically <= 2 words) until
    # we hit the real product name.
    asin_pipe_match = re.search(r'\|\s*B0[A-Z0-9]{8,}', text)
    if asin_pipe_match:
        before_asin = text[:asin_pipe_match.start()]  # "Puma | Hoodie | Black | S"
        segments = [s.strip() for s in before_asin.split('|')]
        # Drop trailing short segments (<=2 words) — these are color/size variants
        while len(segments) > 1 and len(segments[-1].split()) <= 2:
            segments.pop()
        text = ' | '.join(segments)  # rejoin with pipes for readability
    elif '|' in text:
        text = text.split('|')[0]               # fallback: cut at first pipe

    # Remove underscore-separated variant codes that appear AFTER the product name
    text = re.sub(r'\b[A-Z0-9]+(?:_[A-Z0-9]+)+\b', '', text)  # remove ALL_CAPS_CODES
    text = re.sub(r'_+', ' ', text)  # replace any leftover underscores with space

    # Collapse multiple spaces / newlines
    text = re.sub(r'\s+', ' ', text)

    return text.strip().rstrip(',').strip()


def is_credit_note(pdf_path):
    """
    Returns True if the PDF is a Credit Note based on filename pattern.
    Credit notes contain '-C-' in the filename  e.g. 'AMD2-C-894123.pdf'
    Invoices do not,                              e.g. 'AMD2-4483369.pdf'
    """
    filename = os.path.basename(pdf_path)
    return bool(re.search(r'-C-', filename))


def extract_credit_note_data(pdf_path):
    """
    Extracts line items from a Credit Note PDF.
    Returns a list of dicts with keys:
      ASN, Credit Note Date, Credit Note No, Credit Amount, Narration
    """
    data = []

    with pdfplumber.open(pdf_path) as pdf:

        # ── 1. Pull header fields from raw text ──────────────────────────────
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        credit_note_no_match = re.search(r'Credit Note No[:\s]+(\S+)', full_text)
        credit_note_no = credit_note_no_match.group(1) if credit_note_no_match else None

        credit_note_date_match = re.search(r'Credit Note Date[:\s]+(\d{2}\.\d{2}\.\d{4})', full_text)
        credit_note_date = credit_note_date_match.group(1) if credit_note_date_match else None

        # ── 2. Parse line items from tables ──────────────────────────────────
        pending_row = None

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    # Skip rows that don't have enough columns
                    if not row or len(row) < 2:
                        continue

                    # Skip header rows (preserve pending_row across page breaks)
                    if row[0] and row[0].strip().startswith("Sl"):
                        continue

                    desc_cell = row[1] or ""
                    total_cell = row[-1] or ""

                    # Prepend any saved partial description from previous page
                    if pending_row is not None and desc_cell:
                        desc_cell = (pending_row + " " + desc_cell).strip()
                        pending_row = None

                    # Skip summary / footer rows (no ASIN present)
                    if not desc_cell or not re.search(r'B0[A-Z0-9]{8,}', desc_cell):
                        all_price_cells_empty = all(not (row[i] or "").strip() for i in range(2, len(row)))
                        if desc_cell and all_price_cells_empty and not re.search(r'₹', desc_cell):
                            pending_row = desc_cell
                        continue

                    # ── Extract ASIN ──────────────────────────────────────────
                    asin_match = re.search(r'\|\s*(B0[A-Z0-9]{8,})', desc_cell)
                    if not asin_match:
                        asin_match = re.search(r'(B0[A-Z0-9]{8,})', desc_cell)
                    if not asin_match:
                        pending_row = desc_cell
                        continue

                    asin = asin_match.group(1)

                    # ── Extract Credit Amount (may have leading minus: -₹689.00) ──
                    amount_match = re.search(r'-?₹([\d,]+\.\d{2})', total_cell)
                    if not amount_match:
                        pending_row = desc_cell
                        continue

                    # Store as negative to indicate a credit/return
                    credit_amount = -float(amount_match.group(1).replace(",", ""))

                    # ── Clean narration ───────────────────────────────────────
                    desc_clean = re.sub(r'\s*\n\s*', ' ', desc_cell)
                    narration = clean_narration(desc_clean)

                    data.append({
                        "ASN": asin,
                        "Credit Note Date": credit_note_date,
                        "Credit Note No": credit_note_no,
                        "Credit Amount": credit_amount,
                        "Narration": narration,
                    })

                    pending_row = None

    return data


def extract_invoice_data(pdf_path):
    data = []

    with pdfplumber.open(pdf_path) as pdf:

        # ── 1. Pull header fields from raw text ──────────────────────────────
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        invoice_no_match = re.search(r'Invoice Number\s*:\s*(\S+)', full_text)
        invoice_no = invoice_no_match.group(1) if invoice_no_match else None

        invoice_date_match = re.search(r'Invoice Date\s*:\s*(\d{2}\.\d{2}\.\d{4})', full_text)
        invoice_date = invoice_date_match.group(1) if invoice_date_match else None

        # ── 2. Parse line items from tables ──────────────────────────────────
        pending_row = None   # carries a split row that continues on next page

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    # Skip rows that don't have enough columns
                    if not row or len(row) < 2:
                        continue

                    # Skip header rows (but preserve pending_row across page breaks)
                    if row[0] and row[0].strip().startswith("Sl"):
                        continue

                    desc_cell = row[1] or ""
                    total_cell = row[-1] or ""   # last column = Total Amount

                    # Prepend any saved partial description from previous page
                    if pending_row is not None and desc_cell:
                        desc_cell = (pending_row + " " + desc_cell).strip()
                        pending_row = None

                    # Skip summary / footer rows (no ASIN present)
                    if not desc_cell or not re.search(r'B0[A-Z0-9]{8,}', desc_cell):
                        all_price_cells_empty = all(not (row[i] or "").strip() for i in range(2, len(row)))
                        if desc_cell and all_price_cells_empty and not re.search(r'₹', desc_cell):
                            pending_row = desc_cell
                        continue

                    # ── Extract ASIN ──────────────────────────────────────────
                    asin_match = re.search(r'\|\s*(B0[A-Z0-9]{8,})', desc_cell)
                    if not asin_match:
                        asin_match = re.search(r'(B0[A-Z0-9]{8,})', desc_cell)
                    if not asin_match:
                        pending_row = desc_cell
                        continue

                    asin = asin_match.group(1)

                    # ── Extract Total Amount ──────────────────────────────────
                    amount_match = re.search(r'₹([\d,]+\.\d{2})', total_cell)
                    if not amount_match:
                        pending_row = desc_cell
                        continue

                    amount = float(amount_match.group(1).replace(",", ""))

                    # ── Clean narration ───────────────────────────────────────
                    desc_clean = re.sub(r'\s*\n\s*', ' ', desc_cell)
                    narration = clean_narration(desc_clean)

                    data.append({
                        "ASN": asin,
                        "Invoice Date": invoice_date,
                        "Invoice No": invoice_no,
                        "Amount": amount,
                        "Narration": narration,
                    })

                    pending_row = None

    return data


def process_folder(folder_path, output_file="output.json"):
    invoices = []    # list of invoice line items
    credit_notes = []  # list of credit note line items
    inv_sr = 1
    cn_sr = 1

    # Collect all PDFs recursively from folder and all subfolders
    all_pdfs = []
    for root, dirs, files in os.walk(folder_path):
        dirs.sort()   # process subfolders in alphabetical order
        for file in sorted(files):
            if file.lower().endswith(".pdf"):
                all_pdfs.append(os.path.join(root, file))

    for pdf_path in all_pdfs:
        file        = os.path.basename(pdf_path)
        # Show relative path so you can see which subfolder each file came from
        rel_path    = os.path.relpath(pdf_path, folder_path)

        if is_credit_note(pdf_path):
            print(f"Processing [CREDIT NOTE]: {rel_path}")
            extracted = extract_credit_note_data(pdf_path)
            for item in extracted:
                item = {"Sr.No": cn_sr, **item}
                credit_notes.append(item)
                cn_sr += 1
        else:
            print(f"Processing [INVOICE]     : {rel_path}")
            extracted = extract_invoice_data(pdf_path)
            for item in extracted:
                item = {"Sr.No": inv_sr, **item}
                invoices.append(item)
                inv_sr += 1

    # ── Save to separate JSON files ───────────────────────────────────────────
    folder = os.path.dirname(os.path.abspath(output_file))
    invoices_file    = os.path.join(folder, "invoice.json")
    credit_notes_file = os.path.join(folder, "credit_note.json")

    with open(invoices_file, "w", encoding="utf-8") as f:
        json.dump(invoices, f, indent=4, ensure_ascii=False)

    with open(credit_notes_file, "w", encoding="utf-8") as f:
        json.dump(credit_notes, f, indent=4, ensure_ascii=False)

    print(f"\n✅ Done!")
    print(f"   Invoices    : {len(invoices)} line items → {invoices_file}")
    print(f"   Credit Notes: {len(credit_notes)} line items → {credit_notes_file}")

    # ── Create final.json by matching ASNs ───────────────────────────────────
    final_file = os.path.join(folder, "final.json")
    final = create_final_json(invoices, credit_notes, final_file)
    matched   = sum(1 for r in final if r["Credit Note No"] is not None)
    unmatched = sum(1 for r in final if r["Credit Note No"] is None)
    print(f"   Final JSON  : {len(final)} total items ({matched} matched, {unmatched} unmatched) → {final_file}")

    # ── Create final.xlsx with brand-wise sheets ──────────────────────────────
    xlsx_file = os.path.join(folder, "final.xlsx")
    brand_counts = create_final_xlsx(final, xlsx_file)
    print(f"   Final XLSX  : {len(brand_counts)} brand sheets → {xlsx_file}")
    for brand, count in sorted(brand_counts.items()):
        print(f"     • {brand}: {count} rows")

    return invoices, credit_notes


def create_final_json(invoices, credit_notes, output_file="final.json"):
    """
    Merges all invoice line items with credit note data by ASN.
    - Matched ASN  → full merged record
    - Unmatched    → credit fields set to None (null in JSON)
    """
    credit_map = {cn["ASN"]: cn for cn in credit_notes}
    final = []

    for sr, inv in enumerate(invoices, start=1):
        asn = inv["ASN"]
        cn  = credit_map.get(asn)
        final.append({
            "Sr.No":            sr,
            "ASN":              asn,
            "Invoice Date":     inv.get("Invoice Date"),
            "Invoice No":       inv.get("Invoice No"),
            "Amount":           inv.get("Amount"),
            "Narration":        inv.get("Narration"),
            "Credit Note Date": cn.get("Credit Note Date") if cn else None,
            "Credit Note No":   cn.get("Credit Note No")   if cn else None,
            "Credit Amount":    cn.get("Credit Amount")    if cn else None,
        })

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(final, f, indent=4, ensure_ascii=False)

    return final


# ── Brand keywords (ordered: longer/specific first to avoid partial matches) ──
_BRAND_KEYWORDS = [
    # ── Multi-word brands first (must come before their single-word roots) ────
    "Garden Vareli",
    "Van Heusen",
    "Allen Solly",
    "Tommy Hilfiger",
    "Tommy Helfiger",
    "Peter England",
    "Arrow Sports",
    "Calvin Klein",
    "Louis Philippe",
    "U.S. Polo Assn",
    "United Colors of Benetton",
    "Jack & Jones",
    "Being Human",
    "Park Avenue",
    "Monte Carlo",
    "Mast & Harbour",
    "Here & Now",
    "Red Tape",
    "Flying Machine",
    "Indian Terrain",
    "ColorPlus",
    # ── Single-word brands ────────────────────────────────────────────────────
    "Arrow",
    "Raymond",
    "GAP",
    "Vastramay",
    "Puma",
    "PUMA",
    "XYXX",
    "Soch",
    "ALYNE",
    "Levis",
    "Levi's",
    "Wrangler",
    "Pepe",
    "Spykar",
    "Killer",
    "Blackberrys",
    "Zodiac",
    "Turtle",
    "Basics",
    "Breakbounce",
    "Adidas",
    "Reebok",
    "Nike",
    "HRX",
    "Jockey",
    "Rupa",
    "Dollar",
    "Biba",
    "Libas",
    "Rangmanch",
    "Anubhutee",
    "Aurelia",
    "Nayo",
    "Anouk",
    "Vishudh",
    "Tisser",
]

# Canonical name map — normalises casing / spelling variations
_BRAND_CANONICAL = {
    "tommy hilfiger":     "Tommy Hilfiger",
    "tommy helfiger":     "Tommy Hilfiger",
    "arrow sports":       "Arrow Sports",
    "arrow":              "Arrow",
    "levi's":             "Levis",
    "levis":              "Levis",
    "u.s. polo assn":     "US Polo Assn",
    "colorplus":          "ColorPlus",
    "jack & jones":       "Jack & Jones",
    "being human":        "Being Human",
    "flying machine":     "Flying Machine",
    "indian terrain":     "Indian Terrain",
    "monte carlo":        "Monte Carlo",
    "mast & harbour":     "Mast & Harbour",
    "here & now":         "Here & Now",
    "park avenue":        "Park Avenue",
    "red tape":           "Red Tape",
}

def _detect_brand(narration):
    n = (narration or "").lower()
    for brand in _BRAND_KEYWORDS:
        if brand.lower() in n:
            return _BRAND_CANONICAL.get(brand.lower(), brand)
    return "Other"


def create_final_xlsx(final_data, output_file="final.xlsx"):
    """
    Creates a multi-sheet Excel file from final_data (list of dicts).
    Each brand gets its own sheet, named after the brand.
    Null credit fields are highlighted in light yellow.
    Returns a dict of {brand: row_count}.
    """
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    NULL_FILL   = PatternFill("solid", start_color="FFF2CC")
    DATA_FONT   = Font(name="Arial", size=10)
    TOTAL_FILL  = PatternFill("solid", start_color="D9E1F2")
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT       = Alignment(horizontal="right",  vertical="center")
    THIN        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    COLUMNS = [
        ("Sr.No",            8),
        ("ASN",             14),
        ("Invoice Date",    14),
        ("Invoice No",      18),
        ("Amount",          12),
        ("Narration",       45),
        ("Credit Note Date",16),
        ("Credit Note No",  18),
        ("Credit Amount",   14),
    ]
    HEADERS = [c[0] for c in COLUMNS]
    WIDTHS  = [c[1] for c in COLUMNS]

    brand_data = defaultdict(list)
    for row in final_data:
        brand = _detect_brand(row.get("Narration", ""))
        brand_data[brand].append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for brand in sorted(brand_data.keys()):
        rows = brand_data[brand]
        ws   = wb.create_sheet(title=brand[:31])

        for ci, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = CENTER; cell.border = BORDER
        ws.row_dimensions[1].height = 20

        for ri, record in enumerate(rows, 2):
            for ci, key in enumerate(HEADERS, 1):
                value = record.get(key)
                cell  = ws.cell(row=ri, column=ci, value=value)
                cell.font = DATA_FONT; cell.border = BORDER
                if value is None:
                    cell.value = "null"; cell.fill = NULL_FILL
                    cell.alignment = CENTER
                elif key in ("Amount", "Credit Amount"):
                    cell.number_format = '#,##0.00'
                    cell.alignment = RIGHT
                elif key == "Narration":
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER
            ws.row_dimensions[ri].height = 18

        for ci, width in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.freeze_panes = "A2"

        tr = len(rows) + 2
        tc = ws.cell(row=tr, column=1, value="TOTAL")
        tc.font = Font(name="Arial", bold=True, size=10); tc.alignment = CENTER

        for key in ("Amount", "Credit Amount"):
            ci  = HEADERS.index(key) + 1
            col = get_column_letter(ci)
            cell = ws.cell(row=tr, column=ci, value=f"=SUM({col}2:{col}{tr-1})")
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.number_format = '#,##0.00'
            cell.alignment = RIGHT; cell.border = BORDER; cell.fill = TOTAL_FILL
        ws.row_dimensions[tr].height = 18

    wb.save(output_file)
    return {brand: len(rows) for brand, rows in brand_data.items()}


# ── Quick single-file test ────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    if len(sys.argv) == 2 and sys.argv[1].endswith(".pdf"):
        # python main.py some_file.pdf  — auto-detect and process
        pdf_path = sys.argv[1]
        if is_credit_note(pdf_path):
            print("[Detected: CREDIT NOTE]")
            result = extract_credit_note_data(pdf_path)
        else:
            print("[Detected: INVOICE]")
            result = extract_invoice_data(pdf_path)
        print(json.dumps(result, indent=4, ensure_ascii=False))
    else:
        # 🔧 CHANGE THIS PATH to your folder
        folder_path = r"D:\Sumesh\invoice_data_extractor\Jan Amz branded lives\PUMA\New folder"
        process_folder(folder_path)
